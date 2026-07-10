"""把 xiudong.db 里的演出数据导出成一份好看、方便浏览的 Excel 表格。"""
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
DB_PATH = SCRIPT_DIR / "xiudong.db"
OUT_PATH = SCRIPT_DIR / "xiudong_演出数据.xlsx"

COLUMNS = [
    ("title", "演出名称", 45),
    ("performers", "艺人", 20),
    ("city_name", "城市", 10),
    ("site_name", "场馆", 26),
    ("show_time", "演出时间", 16),
    ("price", "票价", 12),
    ("sold_out", "是否售罄", 10),
    ("last_seen_at", "最后更新时间", 18),
]

conn = sqlite3.connect(DB_PATH)
df = pd.read_sql_query(
    "SELECT title, performers, city_name, site_name, show_time, price, sold_out, last_seen_at "
    "FROM shows ORDER BY city_name, show_time", conn)
conn.close()

df["sold_out"] = df["sold_out"].map({1: "是", 0: "否"}).fillna("否")
df.columns = [label for _, label, _ in COLUMNS]

df.to_excel(OUT_PATH, index=False, sheet_name="演出数据")

from openpyxl import load_workbook

wb = load_workbook(OUT_PATH)
ws = wb["演出数据"]

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="4472C4")
body_font = Font(name="Arial")

for col_idx, (_, label, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = width

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical="center", wrap_text=(cell.column <= 2))

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 22

wb.save(OUT_PATH)
print(f"导出完成: {OUT_PATH}，共 {len(df)} 行")
